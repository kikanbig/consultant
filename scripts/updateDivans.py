#!/usr/bin/env python3
"""
Скрипт для обновления базы данных диванов из Excel файла
Автоматически генерирует алиасы для брендов и моделей
"""

import openpyxl
import json
import re
from html import unescape

# Алиасы для брендов
BRAND_ALIASES = {
    'VELUNA': ['veluna', 'велуна', 'велюна', 'илуна', 'iluna', 'вилуна'],
    'ELVA': ['elva', 'элва', 'эльва', 'елва', 'ельва'],
    'Rivalli': ['rivalli', 'ривалли', 'ривали', 'риваллі', 'revolut', 'револют', 'револю'],
    'Мебельград': ['мебельград', 'mebelgrad', 'мебелград'],
    'Mio Tesoro': ['mio tesoro', 'мио тесоро', 'мио тезоро', 'миа тесоро', 'миа тезоро', 'mia tesoro', 'mio', 'миа', 'мио'],
    'Moon Trade': ['moon trade', 'мун трейд', 'мун трэйд', 'moon', 'мун'],
    'Anderssen': ['anderssen', 'андерссен', 'андерсен', 'андерсон'],
    'Moon': ['moon', 'мун', 'мун'],
    'Trade': ['trade', 'трейд', 'трэйд', 'трейт'],
    'Woodcraft': ['woodcraft', 'вудкрафт', 'вуткрафт'],
    'Leset': ['leset', 'лесет', 'лесэт'],
    'Homeme': ['homeme', 'хомми', 'хоумми', 'хомме'],
    'Askona': ['askona', 'аскона'],
    'Lazurit': ['lazurit', 'лазурит'],
    'Pushe': ['pushe', 'пуше', 'пушэ', 'пуш'],
    'First': ['first', 'фирст', 'фёрст', 'ферст']
}

# Словарь транслитерации для моделей
TRANSLIT_MAP = {
    # Специальные случаи (целые слова)
    'yuki': ['юкки', 'юки', 'yukki'],
    'gizela': ['гизела', 'гизелла'],
    'chianti': ['кьянти', 'киянти', 'кианти', 'kyanti'],
    'miami': ['майами', 'маями', 'миами'],
    'aspen': ['аспен', 'аспэн'],
    'leyton': ['лейтон', 'лэйтон'],
    'evas': ['эвас', 'евас'],
    'sonni': ['сонни', 'сони'],
    'eloy': ['элой', 'елой'],
    'vito': ['вито', 'віто'],
    'kubo': ['кубо', 'кубо'],
    'bilbao': ['бильбао', 'билбао', 'бильбао'],
    'pekin': ['пекин', 'пекін', 'beijing'],
    'aisti': ['айсти', 'аисти', 'isti'],
    'riemu': ['риему', 'риэму'],
    'tulisia': ['тулисия', 'тулисія'],
    'saari': ['саари', 'саарі'],
    'unelma': ['унельма', 'унелма'],
    'lintu': ['линту', 'лінту'],
    'lira': ['лира', 'ліра'],
    'tunne': ['тунне', 'туне'],
    'aurinko': ['ауринко', 'аурінко'],
    'velke': ['велке', 'велькэ'],
    'tuuli': ['туули', 'туулі'],
    'toivo': ['тойво', 'тоіво'],
    'jersey': ['джерси', 'джерсі', 'джерси'],
    'emma': ['эмма', 'емма'],
    'dijon': ['дижон', 'діжон'],
    'orleans': ['орлеан', 'орлеан'],
    'parma': ['парма', 'парма'],
    'discovery': ['дискавери', 'дискавері'],
    'porto': ['порто', 'порто'],
    'somerset': ['сомерсет', 'сомерсет'],
    'rimini': ['риммини', 'римини'],
    'valencia': ['валенсия', 'валенсія'],
    'montreal': ['монреаль', 'монреал', 'монтреаль'],
    'douglas': ['дуглас', 'даглас'],
    # Общие буквы (для остальных случаев)
    'a': 'а', 'b': 'б', 'v': 'в', 'g': 'г', 'd': 'д', 'e': 'е',
    'z': 'з', 'i': 'и', 'k': 'к', 'l': 'л', 'm': 'м', 'n': 'н',
    'o': 'о', 'p': 'п', 'r': 'р', 's': 'с', 't': 'т', 'u': 'у',
    'f': 'ф', 'h': 'х', 'w': 'в', 'y': 'й'
}

def generate_phonetic_variants(word):
    """
    Генерирует фонетические варианты произношения слова
    """
    variants = set([word])
    
    # Замены для распространённых фонетических ошибок
    phonetic_rules = [
        ('е', 'э'), ('э', 'е'),  # е/э
        ('и', 'ы'), ('ы', 'и'),  # и/ы
        ('о', 'а'), ('а', 'о'),  # о/а в безударной позиции
        ('ё', 'е'), ('е', 'ё'),  # ё/е
        ('й', 'и'), ('и', 'й'),  # й/и
        ('ц', 'тс'), ('тс', 'ц'),  # ц/тс
        ('ч', 'тш'), ('тш', 'ч'),  # ч/тш
        ('щ', 'шч'), ('шч', 'щ'),  # щ/шч
        ('дж', 'ж'), ('ж', 'дж'),  # дж/ж
        ('нн', 'н'), ('н', 'нн'),  # двойные согласные
        ('лл', 'л'), ('л', 'лл'),
        ('мм', 'м'), ('м', 'мм'),
        ('сс', 'с'), ('с', 'сс'),
        ('тт', 'т'), ('т', 'тт'),
    ]
    
    for old, new in phonetic_rules:
        if old in word:
            variants.add(word.replace(old, new))
    
    return variants

def generate_model_aliases(model_name):
    """
    Генерирует максимально разнообразные алиасы для модели
    """
    if not model_name:
        return []
    
    aliases = set()
    model_lower = model_name.lower().strip()
    
    # Добавляем оригинал
    aliases.add(model_lower)
    
    # Разбиваем на слова
    words = model_lower.split()
    
    # Для каждого слова генерируем фонетические варианты
    for word in words:
        # Фонетические варианты
        for variant in generate_phonetic_variants(word):
            aliases.add(variant)
            
            # Добавляем с остальными словами
            other_words = [w for w in words if w != word]
            if other_words:
                aliases.add(' '.join([variant] + other_words))
                aliases.add(' '.join(other_words + [variant]))
    
    # Для каждого слова проверяем специальные случаи из TRANSLIT_MAP
    for word in words:
        if word in TRANSLIT_MAP and isinstance(TRANSLIT_MAP[word], list):
            # Это специальный случай - добавляем все варианты
            for variant in TRANSLIT_MAP[word]:
                aliases.add(variant)
                
                # Фонетические варианты специальных случаев
                for phonetic in generate_phonetic_variants(variant):
                    aliases.add(phonetic)
                
                # Также добавляем с остальными словами
                other_words = [w for w in words if w != word]
                if other_words:
                    for variant_word in TRANSLIT_MAP[word]:
                        aliases.add(' '.join([variant_word] + other_words))
                        aliases.add(' '.join(other_words + [variant_word]))
    
    # Добавляем первое слово отдельно (основное название)
    if words:
        first_word = words[0]
        aliases.add(first_word)
        
        # Фонетические варианты первого слова
        for variant in generate_phonetic_variants(first_word):
            aliases.add(variant)
        
        # Транслитерация первого слова
        if first_word in TRANSLIT_MAP and isinstance(TRANSLIT_MAP[first_word], list):
            for variant in TRANSLIT_MAP[first_word]:
                aliases.add(variant)
                for phonetic in generate_phonetic_variants(variant):
                    aliases.add(phonetic)
    
    # Убираем суффиксы типа "-4", "-2" для поиска по базовому названию
    for word in words:
        base_word = re.sub(r'-?\d+$', '', word)
        if base_word and base_word != word and len(base_word) >= 3:
            aliases.add(base_word)
            
            # Фонетические варианты базового слова
            for variant in generate_phonetic_variants(base_word):
                aliases.add(variant)
            
            # И транслитерация базового слова
            if base_word in TRANSLIT_MAP and isinstance(TRANSLIT_MAP[base_word], list):
                for variant in TRANSLIT_MAP[base_word]:
                    aliases.add(variant)
                    for phonetic in generate_phonetic_variants(variant):
                        aliases.add(phonetic)
    
    # Ищем латинский ключ для кириллического слова
    for word in words:
        for lat_key, cyr_variants in TRANSLIT_MAP.items():
            if isinstance(cyr_variants, list) and word in cyr_variants:
                aliases.add(lat_key)
                other_words = [w for w in words if w != word]
                if other_words:
                    aliases.add(' '.join([lat_key] + other_words))
                    aliases.add(' '.join(other_words + [lat_key]))
    
    # Убираем слишком короткие алиасы (меньше 3 символов)
    aliases = {a for a in aliases if len(a) >= 3}
    
    return sorted(list(aliases))

def generate_article_aliases(article_code):
    """
    Генерирует алиасы для артикула (код товара)
    Пользователь произносит по одной цифре: "один ноль ноль семь семь один два семь"
    """
    if not article_code:
        return []
    
    # Словарь произношения цифр
    digit_names = {
        '0': ['ноль', 'нуль'],
        '1': ['один', 'раз', 'адин'],
        '2': ['два', 'двойка'],
        '3': ['три', 'тройка'],
        '4': ['четыре', 'четверка', 'читыре'],
        '5': ['пять', 'пятерка', 'пьять'],
        '6': ['шесть', 'шестерка', 'шэсть'],
        '7': ['семь', 'семерка', 'сем'],
        '8': ['восемь', 'восьмерка', 'восем'],
        '9': ['девять', 'девятка', 'дивять']
    }
    
    aliases = set()
    code_str = str(article_code).strip()
    
    # Добавляем оригинал
    aliases.add(code_str)
    
    # Генерируем произношение по цифрам
    # Например: "10077127" → "один ноль ноль семь семь один два семь"
    for i in range(len(code_str)):
        digit = code_str[i]
        if digit in digit_names:
            # Генерируем все комбинации произношений
            # Это будет использоваться для поиска
            pass
    
    # Добавляем варианты с пробелами между цифрами
    spaced = ' '.join(code_str)
    aliases.add(spaced)
    
    return sorted(list(aliases))

def generate_brand_aliases(brand_name):
    """
    Генерирует алиасы для бренда
    """
    if not brand_name:
        return []
    
    brand_clean = brand_name.strip()
    
    # Проверяем точное совпадение
    if brand_clean in BRAND_ALIASES:
        return BRAND_ALIASES[brand_clean]
    
    # Проверяем частичное совпадение
    for key, aliases in BRAND_ALIASES.items():
        if key.lower() in brand_clean.lower() or brand_clean.lower() in key.lower():
            return aliases
    
    # Если не найдено, возвращаем базовые алиасы
    return [brand_clean.lower()]

def parse_excel_to_json(excel_path, output_path):
    """
    Парсит Excel файл и создаёт JSON с алиасами
    """
    print(f"📖 Читаю файл: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Мягкая мебель']
    
    divans = []
    
    # Читаем данные начиная со строки 8
    for row in range(8, ws.max_row + 1):
        kod = ws.cell(row, 1).value  # A - код товара
        name = ws.cell(row, 2).value  # B - название
        description = ws.cell(row, 4).value  # D - общее описание
        
        if kod and name:
            # Очищаем HTML из описания
            if description:
                description = re.sub(r'<[^<]+?>', '', str(description))
                description = unescape(description)
                description = re.sub(r'\s+', ' ', description).strip()
            
            # Улучшенный парсинг бренда и модели
            name_str = str(name)
            
            # Убираем тип мебели в начале
            pattern = r'^(Диван\s+(угловой\s+|П-образный\s+)?|Кресло(-кровать|-реклайнер|\s+мягкое)?\s+|Комплект\s+.*?\s+|Модуль\s+мягкий\s+|Пуф(-трансформер)?\s+|Тахта(\s+угловая)?\s+|Уголок\s+.*?\s+|Скамья\s+.*?\s+|Оттоманка\s+)'
            cleaned_name = re.sub(pattern, '', name_str, flags=re.IGNORECASE).strip()
            
            # Специальная обработка для "Mio Tesoro" (двухсловный бренд)
            brand = ""
            model = ""
            
            if cleaned_name.startswith('Mio Tesoro') or cleaned_name.startswith('Mio tesoro'):
                brand = 'Mio Tesoro'
                model = cleaned_name[10:].strip()  # Убираем "Mio Tesoro"
                # Убираем всё после скобки
                if '(' in model:
                    model = model[:model.index('(')].strip()
            elif cleaned_name.startswith('Moon Trade') or cleaned_name.startswith('Moon trade'):
                brand = 'Moon Trade'
                model = cleaned_name[10:].strip()
                if '(' in model:
                    model = model[:model.index('(')].strip()
            else:
                # Обычный парсинг для односложных брендов
                match = re.search(r'^([A-Za-zА-Яа-я\s]+?)\s+([A-Za-zА-Яа-я0-9\s\-]+?)(?:\s*\(|$)', cleaned_name)
                if match:
                    brand = match.group(1).strip()
                    model = match.group(2).strip()
            
            # Генерируем алиасы
            brand_aliases = generate_brand_aliases(brand)
            model_aliases = generate_model_aliases(model)
            article_aliases = generate_article_aliases(kod)
            
            divans.append({
                'kod': str(kod),
                'name': name_str,
                'brand': brand,
                'model': model,
                'brandAliases': brand_aliases,
                'modelAliases': model_aliases,
                'articleAliases': article_aliases,
                'description': description or 'Описание отсутствует'
            })
    
    # Сохраняем в JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump({'divans': divans}, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Сохранено {len(divans)} диванов в {output_path}")
    
    # Статистика
    print(f"\n📊 Статистика алиасов:")
    total_brand_aliases = sum(len(d['brandAliases']) for d in divans)
    total_model_aliases = sum(len(d['modelAliases']) for d in divans)
    total_article_aliases = sum(len(d['articleAliases']) for d in divans)
    print(f"   Всего алиасов брендов: {total_brand_aliases}")
    print(f"   Всего алиасов моделей: {total_model_aliases}")
    print(f"   Всего алиасов артикулов: {total_article_aliases}")
    
    # Примеры
    print(f"\n📝 Примеры (первые 3):")
    for i, divan in enumerate(divans[:3], 1):
        print(f"\n{i}. {divan['name'][:60]}")
        print(f"   Код: {divan['kod']}")
        print(f"   Бренд: {divan['brand']}")
        print(f"   Алиасы бренда: {', '.join(divan['brandAliases'][:5])}")
        print(f"   Модель: {divan['model']}")
        print(f"   Алиасы модели ({len(divan['modelAliases'])} шт): {', '.join(divan['modelAliases'][:8])}")
        if len(divan['modelAliases']) > 8:
            print(f"      ... и еще {len(divan['modelAliases']) - 8} алиасов")

if __name__ == '__main__':
    import os
    
    # Определяем пути
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)
    excel_path = os.path.join(project_dir, 'Файл для диванов', 'Диваны, крессла, матрасы розница (1).xlsx')
    output_path = os.path.join(project_dir, 'src', 'data', 'divans.json')
    
    print("🚀 Обновление базы данных диванов\n")
    parse_excel_to_json(excel_path, output_path)
    print("\n✨ Готово!")

